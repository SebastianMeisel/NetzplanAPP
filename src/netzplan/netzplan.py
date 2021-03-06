from PIL import Image, ImageDraw, ImageFont  # Netzplan zeichnen und exportieren
import csv  # für CVS-Import
import re
from openpyxl import load_workbook  # für Excel-Import
from openpyxl.utils import get_column_letter  # Spalten-Namen in Excel
from typing import List, Dict

# Activate Logging
import logging

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler("netzplan.log")
logger.addHandler(file_handler)
stream_handler = logging.StreamHandler()
logger.addHandler(stream_handler)


# Netzplan berechnen und zeichnen
version = 0.2


#######################################################################################
# Projekt-Objekt


class Projekt(object):

    # Konstruktor
    def __init__(self, ID: int, Bezeichnung: str):
        self.ID = ID
        self.Bezeichnung = Bezeichnung
        ##############
        self.ArbeitsPakete: Dict[str, ArbeitsPaket] = {}
        self.KritischerPfad: List[str] = []
        self.AP_ID = 0  # Arbeitspacket-Identifier automatisch hochzählen
        self.Ressourcen: Dict[str, Ressource] = {}

    # Arbeitspaket hinzufügen
    def NeuesArbeitsPaket(self, Bezeichnung: str, PT: int, ID=None):
        self.AP_ID += 1
        AP = ArbeitsPaket(ID if ID else self.AP_ID, Bezeichnung, PT, self)
        self.ArbeitsPakete[str(AP.ID)] = AP

    # Ressource hinzufügen
    def NeueRessource(self, ID: str, Name: str):
        R = Ressource(Name, self)
        self.Ressourcen[ID] = R

    # Ressource zuweisen
    def RessourceZuweisen(self, RessourcenID: str, ArbeitsPaketID: str, Kapazität=100):
        self.Ressourcen[RessourcenID].NeuesArbeitsPaket(ArbeitsPaketID, Kapazität)

    # Arbeispaketlist als CSV importieren
    def ImportiereArbeitsPaketListeVonCSV(self, Dateiname: str):
        with open(Dateiname, newline="") as csvfile:
            CSV = csv.DictReader(csvfile, delimiter=";", quotechar='"')
            for Zeile in CSV:
                self.NeuesArbeitsPaket(
                    Zeile["Beschreibung"], int(Zeile["Dauer"]), Zeile["ID"]
                )
                Folgt = re.sub(r"\s+", "", Zeile["Folgt"]).split(",")
                print(Folgt)
                if len(Folgt) == 1 and not Folgt[0] == "":
                    self.ArbeitsPakete[Zeile["ID"]].Folgt(Zeile["Folgt"])
                elif Folgt[0] != "":
                    self.ArbeitsPakete[Zeile["ID"]].Folgt(Folgt)

    # Ressourcen als CSV importieren
    def ImportiereRessourcenVonCSV(self, Dateiname: str):
        with open(Dateiname, newline="") as csvfile:
            CSV = csv.DictReader(csvfile, delimiter=";", quotechar='"')
            for Zeile in CSV:
                R_ID = Zeile["ID"]
                Name = "{VN} {NN}".format(VN=Zeile["Vorname"], NN=Zeile["Nachname"])
                self.NeueRessource(R_ID, Name)
                for AP in Zeile["Arbeitspakete"].split(","):
                    ID_K = AP.split(":", 1)  # in ID und Kapazität aufspalten
                    AP_ID = ID_K[0]  # ID
                    K = 100 if len(ID_K) == 1 else int(ID_K[1])  # Kapazität
                    self.RessourceZuweisen(R_ID, AP_ID, K)

    # Projekt aus Excel importieren
    def ImportiereVonExcel(self, Dateiname: str) -> str:
        Workbook = load_workbook(filename=Dateiname)

        def SpaltenVonTabelle(Tabelle):
            return {
                cell.value: {
                    "Buchstabe": get_column_letter(cell.column),
                    "Nummer": cell.column - 1,
                }
                for cell in Tabelle[1]
                if cell.value
            }

        # Projekt
        if "Projekt" not in Workbook.sheetnames:  # Tabelle Projekt darf NICHT fehlen!
            logger.warning("Tabelle 'Projekt' fehlt oder hat den falschen Namen.")
            return "Tabelle 'Projekt' fehlt oder hat den falschen Namen."

        Tabelle = Workbook["Projekt"]  # Tabelle einlesen
        Spalten = SpaltenVonTabelle(Tabelle) or []  # Spalten einlesen

        # Pflichtspalten überprüfen
        for Spalte in ["ID", "Beschreibung", "Dauer", "Folgt"]:
            if not Spalte in Spalten:
                logger.warning(
                    f"Spalte '{Spalte}' fehlt in der Tabelle '{Tabelle.title}'!"
                )
                return f"Spalte '{Spalte}' fehlt in der Tabelle '{Tabelle.title}'!"

        for AP, row in enumerate(Tabelle.rows):
            if AP > 0:
                ID = Tabelle[Spalten["ID"]["Buchstabe"]][AP].value or ""
                if not type(ID) == str:
                    ID = str(ID)
                Beschreibung = Tabelle[Spalten["Beschreibung"]["Buchstabe"]][AP].value
                Dauer = Tabelle[Spalten["Dauer"]["Buchstabe"]][AP].value
                if Dauer == 0:  # Dauer darf nicht 0 sein.
                    if Beschreibung == "":  # leere Zeile
                        break  # Verarbeitung der Tabelle beenden
                    logger.warning(f"Dauer für ist nicht gesetz!")
                    return f"Dauer für {Beschreibung} ({ID}) ist nicht gesetz!"
                Folgt = Tabelle[Spalten["Folgt"]["Buchstabe"]][AP].value or ""
                # Leerzeichen entfernen und Int in String umwandeln
                Folgt = Folgt.replace(" ", "") if type(Folgt) == str else str(Folgt)
                #
                self.NeuesArbeitsPaket(Beschreibung, int(Dauer), ID)
                # Vorgänger in Liste aufteilen
                if len(Folgt.split(",")) == 1 and not Folgt.split(",")[0] == "":
                    # Überprüfen ob Vorgänger existiert, sonst Fehler melden
                    if Folgt in self.ArbeitsPakete:
                        self.ArbeitsPakete[ID].Folgt(Folgt)
                    else:
                        logger.warning(f"Vorgänger-ID existiert nicht.")
                        return f"ID {Folgt} wird in der Tabelle Projekt als Vorgänger genannt. Sie existiert aber nicht."
                elif Folgt.split(",")[0] != "":
                    for Vorgänger in Folgt.split(","):
                        # Überprüfen ob Vorgänger existiert, sonst Fehler melden
                        if Vorgänger in self.ArbeitsPakete:
                            self.ArbeitsPakete[ID].Folgt(Vorgänger)
                        else:
                            logger.warning(f"Vorgänger-ID existiert nicht.")
                            return f"ID {Vorgänger} wird in der Tabelle Projekt als Vorgänger genannt. Sie existiert aber nicht."

        # Ressourcen
        if "Ressourcen" in Workbook.sheetnames:  # Tabelle Ressourcen darf fehlen
            Tabelle = Workbook["Ressourcen"]
            Spalten = SpaltenVonTabelle(Tabelle) if type(Tabelle) is not None else []

            # Pflichtspalten überprüfen
            for Spalte in ["ID", "Vorname", "Nachname", "Arbeitspakete"]:
                if not Spalte in Spalten:
                    logger.warning(
                        f"Spalte '{Spalte}' fehlt in der Tabelle '{Tabelle.title}'!"
                    )
                    return f"Spalte '{Spalte}' fehlt in der Tabelle '{Tabelle.title}'!"

            for R, row in enumerate(Tabelle.rows):
                if R > 0:
                    R_ID = Tabelle[Spalten["ID"]["Buchstabe"]][R].value or ""
                    Name = (
                        "{VN} {NN}".format(
                            VN=Tabelle[Spalten["Vorname"]["Buchstabe"]][R].value,
                            NN=Tabelle[Spalten["Nachname"]["Buchstabe"]][R].value,
                        )
                        or ""
                    )
                    self.NeueRessource(R_ID, Name)
                    Zeilen: str
                    Zeilen = (
                        Tabelle[Spalten["Arbeitspakete"]["Buchstabe"]][R].value or ""
                    )
                    Zeile: str
                    for Zeile in Zeilen.split(","):
                        Zeile = Zeile.replace(" ", "")  # Leerzeichen entfernen
                        ID_K = Zeile.split(":", 1)  # in ID und Kapazität aufspalten
                        AP_ID = ID_K[0]  # Arbeitspacket-ID
                        K = 100 if len(ID_K) == 1 else int(ID_K[1])  # Kapazität
                        self.RessourceZuweisen(R_ID, AP_ID, K)
        logger.info("Import erfolgreich")
        return ""

    # Vorwärts- und rückwarts-rechnen
    def DurchRechnen(self):
        # Hilfsfunktionen
        def VorwärtsRechnen(AP: ArbeitsPaket):
            AP.getFXZ()
            for NF in AP.Nachfolger:
                VorwärtsRechnen(NF[0])

        def RückwärtsRechnen(AP: ArbeitsPaket):
            AP.getSXZ()
            if AP.GP == 0 and not AP.ID in self.KritischerPfad:
                self.KritischerPfad.append(AP.ID)
            for VG in AP.Vorgänger:
                for i, NF in enumerate(VG.Nachfolger):
                    if type(NF[0]) is not tuple:
                        if NF[0].ID == AP.ID:
                            VG.Nachfolger[i] = (
                                AP,
                                1
                                if len(AP.Nachfolger) == 0
                                else max(t[-1] for t in AP.Nachfolger) + 1,
                            )
                # Wenn Vorgänger noch nicht berechnet ist dort weitermachen
                RückwärtsRechnen(VG)

        # Kapazität je Arbeitspacket berechnen -> Dauer berechnen
        for AP in list(self.ArbeitsPakete.values()):
            PersonenKapazität: float
            PersonenKapazität = 0  # Personen * Kapazität%
            # Personen-Kapazität berechnen
            for R in AP.Ressourcen:
                PersonenKapazität += R.ArbeitsPakete[AP.ID] / 100
            # Wenn keine Resourchen zugeordnet sind, dann mit einer Person, 100% rechnen
            if PersonenKapazität == 0:
                PersonenKapazität = 1
            # Dauer = PersonenTage / PersonenKapazität
            AP.Dauer = int(AP.PT / PersonenKapazität) + (
                AP.PT % PersonenKapazität > 0
            )  # Aufrunden
        # Vorwärts- und Rückwärtsrechnen
        AP = list(self.ArbeitsPakete.values())[0]
        VorwärtsRechnen(AP)
        for ap in reversed(list(self.ArbeitsPakete.values())):
            if len(ap.Nachfolger) == 0:
                AP = ap
                break
        RückwärtsRechnen(AP)

    # Kritischen Pfad ausgeben
    def ZeigeKritischenPfad(self):
        self.DurchRechnen()
        print("Kritischer Pfad: [ ", end="")
        for i, AP_ID in enumerate(reversed(self.KritischerPfad)):
            print(AP_ID, end=" ")
            if i < len(self.KritischerPfad) - 1:
                print(" - ", end="")
        print("]")


###############################################################
# Arbeitspacket-Objekt
class ArbeitsPaket(object):

    # Konstruktor #############################################
    def __init__(self, ID, Bezeichnung: str, PT: int, Projekt: Projekt):
        self.ID = ID
        self.Bezeichnung = Bezeichnung
        self.PT = PT  # Personentage
        self.Projekt = Projekt  # einem Projekt zuordnen
        ###############################
        self.Dauer = self.PT
        self.FAZ = 0  # Früheste Anfangszeit
        self.FEZ = self.Dauer  # Früheste Endzeit
        self.SAZ = 0  # Späteste Anfangszeit
        self.SEZ = 0  # Späteste Endzeit
        self.GP = 0  # Gesamtpuffer
        self.FP = 0  # Freier Puffer
        self.Nachfolger: List[list] = []  # Liste der Nachfolger
        self.Vorgänger: List[ArbeitsPaket] = []  # Liste der Vorgänger
        self.Knoten = None  # Knoten im Netzplan
        self.Ressourcen: List[Ressource] = []

    # Vorgänger hinzufügen
    def Folgt(self, Vorgänger):
        # Unterscheide ob einzelnes Arbeitspacket oder Liste
        if type(Vorgänger) is list:
            for V in Vorgänger:
                self.Vorgänger.append(self.Projekt.ArbeitsPakete[V])
                self.Projekt.ArbeitsPakete[V].Nachfolger.append(
                    [self, 1]
                )  # Zum Vorgänger als Nachfolger hinzufügen + Zähler für Nachfolger vorbereiten
        else:
            self.Vorgänger.append(
                self.Projekt.ArbeitsPakete[Vorgänger]
            )  # Vorgänger hinzufügen
            self.Projekt.ArbeitsPakete[Vorgänger].Nachfolger.append(
                [self, 1]
            )  # Zum Vorgänger als Nachfolger hinzufügen

    # Früheste Anfangs- und Endzeit bestimmen
    def getFXZ(self):
        # Früheste Anfangszeit
        if len(self.Vorgänger) > 0:
            self.FAZ = max(AP.FEZ for AP in self.Vorgänger)
        # Frühester Endzeitpunkt
        self.FEZ = self.Dauer + self.FAZ

    # Späteste Anfangs- und Endzeit und Puffer bestimmen
    def getSXZ(self):
        # Späteste Endzeit
        if len(self.Nachfolger) > 0:
            self.SEZ = min(AP[0].SAZ for AP in self.Nachfolger)
        else:  # Wenn kein Nachfolger nach früheste Endzeit übernehmen
            self.SEZ = self.FEZ

        # Späteste Anfangszeit
        self.SAZ = self.SEZ - self.Dauer

        # Gesamtpuffer
        self.GP = self.SEZ - self.FEZ

        # Freier Puffer
        if len(self.Nachfolger) > 0:
            self.FP = min(AP[0].FAZ for AP in self.Nachfolger) - self.FEZ


#######################################################################################
# Resource-Objekt


class Ressource(object):
    # Konstrukor
    def __init__(self, Name: str, Projekt: Projekt):
        self.Name = Name
        self.Projekt = Projekt  # Projekt zuordnen
        #################
        self.ArbeitsPakete: Dict[
            ArbeitsPaket, int
        ] = {}  # Arbeitspackete und Kapazität, die der Ressource zugeordnet werden

    def NeuesArbeitsPaket(self, AP: str, Kapazität=100):
        self.ArbeitsPakete[self.Projekt.ArbeitsPakete[AP].ID] = Kapazität
        self.Projekt.ArbeitsPakete[AP].Ressourcen.append(self)


####################################################################
# Netzplan-Object
class Netzplan(object):

    # Kontruktor
    def __init__(self, Name: str):
        self.Name = Name  # Name der Exportdatei
        # A4-Seite als Leinwand
        self.x = 3508  # X-Länge der Leinwand
        self.y = 2480  # Y-Länge der Leinwand
        self.a4image = Image.new(
            "RGB", (self.x, self.y), (255, 255, 255)  # A4 bei 72dpi
        )  # Weiß
        # Schriftart
        self.font = ImageFont.truetype("SourceCodePro-Light.ttf", 24, 0)
        self.bold_font = ImageFont.truetype("SourceCodePro-Bold.ttf", 24)
        self.heading_font = ImageFont.truetype("SourceCodePro-Bold.ttf", 36)

        # Zeichnung um Netzplan aufzunehmen
        self.Zeichnung = ImageDraw.Draw(self.a4image)
        # Listen: Knoten Raster
        self.Knoten: List[int] = []  # Knoten auf der Zeichnung
        self.Raster: List[
            str
        ] = (
            []
        )  # Liste der Belegten Positionen im Raster um Überschneidungen zu vermeiden

    # Knoten hinzufügen
    def NeuerKnoten(self, x: float, y: float, AP: ArbeitsPaket):
        # Knoten-Objekt …
        K: Knoten = Knoten(AP.ID, x, y, AP, self.Zeichnung)  # … anlegen
        K.Zeichnen()  # … zeichnen
        self.Knoten.append(K.ID)  # … (ID) in Knoten-Liste des Netzplans eintragen
        self.Raster.append(str(x) + str(y))
        AP.Knoten = K  # … dem ArbeitsPaket zuordnen

    # Netzplan zeichnen
    def Zeichnen(self, Projekt: Projekt):
        x = 0.5
        y = 0.5
        Projekt.DurchRechnen()
        AP: ArbeitsPaket = list(Projekt.ArbeitsPakete.values())[0]
        self.NeuerKnoten(x, y, AP)
        # Hilfsfunktion
        def NachfolgerZeichnen(x: float, y: float, AP: ArbeitsPaket):
            x += 1
            # Wenn Rasterpunkt belegt, dann neue Zeile anfangen.
            for NF in reversed(sorted(AP.Nachfolger, key=lambda liste: liste[-1])):
                if NF[0].ID not in self.Knoten:
                    y += 1
                    while str(x) + str(y) in self.Raster:
                        y += 1
                    self.NeuerKnoten(x, y, NF[0])
                    # Nachfolger zeichnen
                    NachfolgerZeichnen(x, y - 1, NF[0])
                # Verbinder Zeichnen
                fill = (
                    255 if AP.GP == 0 and NF[0].GP == 0 else 0,
                    0,
                    0,
                )  # rot für kritischen Pfad
                width = (
                    3 if AP.GP == 0 and NF[0].GP == 0 else 1
                )  # fett für kritischen Pfad
                xa, ya = AP.Knoten.aus  # Startpunkt des Verbinders
                xb, yb = NF[0].Knoten.ein  # Endpunkt des Verbinders
                ux = AP.Knoten.ux  # X-Größe eines Kästchens
                uy = AP.Knoten.uy  # X-Größe eines Kästchens
                ry = (
                    0 if yb - ya == 0 else 1 if yb - ya > 0 else -1
                )  # Hoch/runter/geradeaus
                rx = 1 if xb - xa >= 0 else 0  # Links/rechts
                xm = (xa + xb) / 2  # Mitte (X-Achse) zwischen Start- und Endpunkt
                # 1. Strich
                # temporäre x und y Werte
                t_xa = xa
                t_ya = ya
                t_xb = xa + ux + rx
                t_yb = ya + (ry * 3 * uy) - 10 * ry
                self.Zeichnung.line((t_xa, t_ya, t_xb, t_yb), fill=fill, width=width)
                # 2. Strich
                t_xa = t_xb
                t_ya = t_yb
                if t_xa <= xb - 2 * (rx * ux):
                    t_xb = xb - 2 * (rx * ux)
                    # t_yb = ya+(ry*3*uy)-10*ry
                    self.Zeichnung.line(
                        (t_xa, t_ya, t_xb, t_yb), fill=fill, width=width
                    )
                # 3. Strich
                t_xa = t_xb
                # t_ya = t_yb
                t_xb = xb - ux + 1
                t_yb = ya + (ry * 6 * uy) - 10 * ry
                self.Zeichnung.line((t_xa, t_ya, t_xb, t_yb), fill=fill, width=width)
                # 4. Strich
                if t_yb != yb - 10 * ry:
                    t_xa = t_xb
                    t_ya = t_yb
                    # t_xb = xb-ux
                    t_yb = yb - 10 * ry
                    self.Zeichnung.line(
                        (t_xa, t_ya, t_xb, t_yb), fill=fill, width=width
                    )
                # 5. Strich
                t_xa = t_xb
                t_ya = t_yb
                t_xb = xb
                # t_yb = yb-10*ry
                self.Zeichnung.line((t_xa, t_ya, t_xb, t_yb), fill=fill, width=width)

        NachfolgerZeichnen(x, y - 1, AP)
        ##########################################
        # Legende
        L = Legende()
        self.NeuerKnoten(8.5, 11.25, L)
        x = 10.5 * AP.Knoten.dx - (30 * 12)
        y = 11.25 * AP.Knoten.dy - AP.Knoten.uy
        for Label, Erklärung in [
            ["ID", "Identifier"],
            ["D", "Dauer"],
            ["FAZ/FEZ", "Früheste Anfangs-, bzw. Endzeit"],
            ["SAZ/SEZ", "Späteste Anfangs-, bzw. Endzeit"],
            ["GP/FP", "Gesamt-, bzw. Freier Puffer"],
        ]:
            y += 30
            self.Zeichnung.text(
                (x, y),
                "{Label:<10}: {Erklärung:<20}".format(Label=Label, Erklärung=Erklärung),
                (0, 0, 0),
                font=self.font,
            )
        ##########################################
        # Arbeitspacket-Liste
        y = (
            self.y - AP.Knoten.dy - (30 * (len(self.Knoten) + 1)) - 35
        )  # Unterer Seitenrand - 30px pro Zeile - 35px für Projektname
        x = AP.Knoten.dx  # Auf X-Achse am ersten Knoten ausrichten
        R = (
            "Ressourcen" if len(Projekt.Ressourcen) > 0 else ""
        )  # Spalte Ressourcen nur, wenn Ressourcen geplant
        self.Zeichnung.text(
            (x, y),
            "Projekt: {Name:<60}".format(Name=Projekt.Bezeichnung),
            (0, 0, 0),
            font=self.heading_font,
        )
        y += 40
        self.Zeichnung.text(
            (x, y),
            "ID {A:<6}: {B:<25}: {C:^7}: {D:<40}".format(
                A="", B="Bezeichnung", C="Dauer", D=R
            ),
            (0, 0, 0),
            font=self.bold_font,
        )
        for AP in list(Projekt.ArbeitsPakete.values()):
            y += 30
            # Ressourcen checken
            Ressourcen = ""
            i = 0  # zähle Ressourcen des Arbeitspackets
            for R in list(Projekt.Ressourcen.values()):
                if str(AP.ID) in R.ArbeitsPakete.keys():
                    i += 1
                    if i > 1:
                        Ressourcen += ", "
                    Ressourcen += R.Name
                    if R.ArbeitsPakete[str(AP.ID)] != 100:
                        Ressourcen += "(" + str(R.ArbeitsPakete[str(AP.ID)]) + "%)"
            self.Zeichnung.text(
                (x, y),
                "AP {ID:<6}: {Bezeichnung:<25}: {Dauer:7}: {Ressourcen:<40} ".format(
                    ID=AP.ID,
                    Bezeichnung=AP.Bezeichnung,
                    Dauer=AP.Dauer,
                    Ressourcen=Ressourcen,
                ),
                (0, 0, 0),
                font=self.font,
            )

    # PDF-Export
    def PdfExport(self, Pfad="."):
        ## als PDF speichern
        self.a4image.save(Pfad + "/" + self.Name + ".pdf", "PDF", dpi=(300, 300))

    # JPG-Export
    def JPGExport(self, Pfad="."):
        ## als JPG speichern
        self.a4image.save(Pfad + "/" + self.Name + ".jpg", dpi=(300, 300))


################################################################
# Pseudo-Object für Legende
class Legende(object):
    ID = "ID"
    Bezeichnung = "Bezeichnung"
    ###############################
    Dauer = "D"
    FAZ = "FAZ"  # Früheste Anfangszeit
    FEZ = "FEZ"  # Früheste Endzeit
    SAZ = "SAZ"  # Späteste Anfangszeit
    SEZ = "SEZ"  # Späteste Endzeit
    GP = "GP"  # Gesamtpuffer
    FP = "FP"  # Freier Puffer


#################################################################
# Knoten-Object
class Knoten(object):
    def __init__(
        self, ID: int, x: float, y: float, AP: ArbeitsPaket, Zeichnung: object
    ):
        self.ID = ID  # ID des Knotens
        self.x = x  # X im Knotenraster
        self.y = y  # Y im Knotenraster
        self.AP = AP  # Arbeitspacket, das der Knoten darstellt
        self.Zeichnung = (
            Zeichnung  # Zeichnungs-Objekt, auf dem der Knoten dargestellt wird
        )
        ##
        # Einheiten für das Zeichnen
        self.ux = 60  # Breite eines Kästchens 14px ~ 0,75cm
        self.uy = 30  # Höhe eines Kästchens 14px ~ 0,50cm
        self.dx = self.ux * 3 + self.ux * 2  # X-Raster: ~2.25cm/Knoten + ~1cm Abstand
        self.dy = self.uy * 4 + self.uy * 2  # Y-Raster: ~2cm/Knoten + ~1cm Abstand
        # Eingang- / Ausgang für Verbinder
        self.ein = (self.x * self.dx, self.y * self.dy + 2 * self.uy)
        self.aus = (self.x * self.dx + 3 * self.ux, self.y * self.dy + 2 * self.uy)
        # Schriftart
        self.font = ImageFont.truetype("SourceCodePro-Light.ttf", 24)
        self.bold_font = ImageFont.truetype("SourceCodePro-Bold.ttf", 24)

    # Knoten zeichnen
    def Zeichnen(self):
        # FAZ
        xa = self.dx * self.x
        ya = self.dy * self.y
        xb = xa + self.ux
        yb = ya + self.uy
        self.Zeichnung.rectangle(
            (xa, ya, xb, yb), fill=(255, 255, 255), outline=(0, 0, 0, 0)
        )
        self.Zeichnung.text(
            (xa + 2, ya + 1),
            "{a:^4}".format(a=str(self.AP.FAZ)),
            (0, 0, 0),
            font=self.font,
        )
        # FEZ
        xa = xb + self.ux
        # ya = ya
        xb = xa + self.ux
        # yb = ya + self.uy
        self.Zeichnung.rectangle(
            (xa, ya, xb, yb), fill=(255, 255, 255), outline=(0, 0, 0, 0)
        )
        self.Zeichnung.text(
            (xa + 2, ya + 1),
            "{a:^4}".format(a=str(self.AP.FEZ)),
            (0, 0, 0),
            font=self.font,
        )
        # ID
        xa = self.dx * self.x
        ya = yb
        xb = xa + 3 * self.ux
        yb = ya + self.uy
        self.Zeichnung.rectangle(
            (xa, ya, xb, yb), fill=(255, 255, 255), outline=(0, 0, 0, 0)
        )
        self.Zeichnung.text(
            (xa + 2, ya + 1),
            "{a:^12}".format(a=str(self.AP.ID)),
            (0, 0, 0),
            font=self.font,
        )
        # Dauer
        xa = self.dx * self.x
        ya = yb
        xb = xa + self.ux
        yb = ya + self.uy
        self.Zeichnung.rectangle(
            (xa, ya, xb, yb), fill=(255, 255, 255), outline=(0, 0, 0, 0)
        )
        self.Zeichnung.text(
            (xa + 2, ya + 1),
            "{a:^4}".format(a=str(self.AP.Dauer)),
            (0, 0, 0),
            font=self.font,
        )
        # GP
        xa = xa + self.ux
        # ya = ya
        xb = xa + self.ux
        # yb = ya + self.uy
        self.Zeichnung.rectangle(
            (xa, ya, xb, yb), fill=(255, 255, 255), outline=(0, 0, 0, 0)
        )
        self.Zeichnung.text(
            (xa + 2, ya + 1),
            "{a:^4}".format(a=str(self.AP.GP)),
            (255 if self.AP.GP == 0 else 0, 0, 0),
            font=self.bold_font,
        )
        # FP
        xa = xa + self.ux
        # ya = ya
        xb = xa + self.ux
        # yb = ya + self.uy
        self.Zeichnung.rectangle(
            (xa, ya, xb, yb), fill=(255, 255, 255), outline=(0, 0, 0, 0)
        )
        self.Zeichnung.text(
            (xa + 2, ya + 1),
            "{a:^4}".format(a=str(self.AP.FP)),
            (0, 0, 0),
            font=self.font,
        )
        # SAZ
        xa = self.dx * self.x
        ya = yb
        xb = xa + self.ux
        yb = ya + self.uy
        self.Zeichnung.rectangle(
            (xa, ya, xb, yb), fill=(255, 255, 255), outline=(0, 0, 0, 0)
        )
        self.Zeichnung.text(
            (xa + 2, ya + 1),
            "{a:^4}".format(a=str(self.AP.SAZ)),
            (0, 0, 0),
            font=self.font,
        )
        # SEZ
        xa = xb + self.ux
        # ya = ya
        xb = xa + self.ux
        # yb = ya + self.uy
        self.Zeichnung.rectangle(
            (xa, ya, xb, yb), fill=(255, 255, 255), outline=(0, 0, 0, 0)
        )
        self.Zeichnung.text(
            (xa + 2, ya + 1),
            "{a:^4}".format(a=str(self.AP.SEZ)),
            (0, 0, 0),
            font=self.font,
        )
